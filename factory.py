import base64
import json
import logging
import os
import re
import zipfile
from io import BytesIO

import requests
from dotenv import load_dotenv
from flask import Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("factory-scanner")

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
PUBLIC_DIR = os.path.join(BASE_DIR, "public")
app = Flask(__name__, static_folder=PUBLIC_DIR, static_url_path="/_static")
CORS(app, resources={r"/api/*": {"origins": "*"}, r"/*": {"origins": "*"}})

app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024

# ── LLM PROVIDERS (OpenAI-compatible; multi-provider fallback) ──────────────────
# Attempts run in the order providers are listed, then by each provider's model
# chain. On a 429 (shared free pools throttle upstream) or failure, we advance to
# the next (provider, model) pair. Configure each provider purely via env; a
# provider is skipped entirely if its API key is absent.
LLM_MAX_TOKENS = int(os.environ.get("LLM_MAX_TOKENS", "3000"))


def _models(env_val: str, default: str) -> list[str]:
    return [m.strip() for m in os.environ.get(env_val, default).split(",") if m.strip()]


def _build_providers() -> list[dict]:
    providers = []
    # Cerebras — dedicated silicon, no shared-pool throttling; primary by default.
    if os.environ.get("CEREBRAS_API_KEY"):
        providers.append({
            "name": "cerebras",
            "url": os.environ.get("CEREBRAS_API_URL", "https://api.cerebras.ai/v1/chat/completions"),
            "key": os.environ["CEREBRAS_API_KEY"].strip(),
            "models": _models("CEREBRAS_MODELS", "gemma-4-31b"),
        })
    # OpenRouter — free multilingual vision pool; good fallback.
    if os.environ.get("OPENROUTER_API_KEY"):
        providers.append({
            "name": "openrouter",
            "url": os.environ.get("OPENROUTER_API_URL", "https://openrouter.ai/api/v1/chat/completions"),
            "key": os.environ["OPENROUTER_API_KEY"].strip(),
            "models": _models(
                "OPENROUTER_MODELS",
                "google/gemma-4-31b-it:free,"
                "google/gemma-4-26b-a4b-it:free,"
                "nvidia/nemotron-3-nano-omni-30b-a3b-reasoning:free",
            ),
        })
    # Groq — legacy fallback (kept for compatibility).
    if os.environ.get("GROQ_API_KEY"):
        providers.append({
            "name": "groq",
            "url": os.environ.get("GROQ_API_URL", "https://api.groq.com/openai/v1/chat/completions"),
            "key": os.environ["GROQ_API_KEY"].strip(),
            "models": _models("GROQ_MODELS", "qwen/qwen3.6-27b"),
        })
    return providers


PROVIDERS = _build_providers()
# Flat ordered list of (provider, model) attempts.
ATTEMPTS  = [(p, m) for p in PROVIDERS for m in p["models"]]

if not PROVIDERS:
    log.warning("No provider API keys set (CEREBRAS_API_KEY / OPENROUTER_API_KEY / GROQ_API_KEY) — /extract will return 500.")
else:
    log.info("Providers active: %s", ", ".join(f"{p['name']}({len(p['models'])})" for p in PROVIDERS))


# ── OCR PROMPT ────────────────────────────────────────────────────────────────

OCR_PROMPT = """\
Extract ALL content from this image into a JSON table preserving the exact grid layout.
Output ONLY this JSON, nothing else:
{"table":[["તારીખ","35","36",...],["1",34,null,...],["2",null,35,...],...]}

RULES:
1. Include EVERY row and column visible — headers, labels, data, blank rows — nothing skipped.
2. Numbers → JSON number (34, not "34"). Text (including Gujarati ગુજરાતી script) → JSON string exactly as written. Blank/empty/illegible → null.
3. Gujarati and other non-Latin scripts: copy the exact Unicode characters as-is into the JSON string.
4. All rows must have the same number of elements (pad with null on the right).
5. Handwriting: 1↔7, 0↔6, 3↔8, 4↔9 are common confusions — read carefully.
6. Zero (0) is valid, never replace with null.
7. Output ONLY the raw JSON object, no markdown, no explanation.\
"""


# ── IMAGE PREPROCESSING ───────────────────────────────────────────────────────

def preprocess_image(image_b64: str) -> tuple[str, str]:
    try:
        from PIL import Image, ImageEnhance
        raw = base64.b64decode(image_b64)
        img = Image.open(BytesIO(raw)).convert("RGB")
        max_dim = max(img.width, img.height)
        if max_dim < 1400:
            scale = 1400 / max_dim
            img = img.resize(
                (min(int(img.width * scale), 2800), min(int(img.height * scale), 2800)),
                Image.LANCZOS,
            )
        elif max_dim > 2800:
            scale = 2800 / max_dim
            img = img.resize((int(img.width * scale), int(img.height * scale)), Image.LANCZOS)
        img = ImageEnhance.Contrast(img).enhance(1.6)
        img = ImageEnhance.Sharpness(img).enhance(2.8)
        img = ImageEnhance.Brightness(img).enhance(1.08)
        buf = BytesIO()
        img.save(buf, format="JPEG", quality=92, optimize=True)
        buf.seek(0)
        return base64.b64encode(buf.getvalue()).decode(), "image/jpeg"
    except Exception as exc:
        log.warning("Preprocessing failed (%s); using original.", exc)
        return image_b64, "image/jpeg"


# ── LLM API (OpenAI-compatible) ─────────────────────────────────────────────────

def call_llm(image_b64: str, media_type: str, provider: dict, model: str) -> requests.Response:
    payload = {
        "model": model,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "text", "text": OCR_PROMPT},
                {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{image_b64}"}},
            ],
        }],
        "temperature": 0,
        "response_format": {"type": "json_object"},
    }
    # Explicit cap when configured (>0). Omitting makes some providers default to
    # ~256 tokens and truncate large tables — keep this set.
    if LLM_MAX_TOKENS > 0:
        payload["max_tokens"] = LLM_MAX_TOKENS  # OpenAI-standard; Groq/OpenRouter/Cerebras
    headers = {
        "Authorization": f"Bearer {provider['key']}",
        "Content-Type": "application/json",
    }
    # OpenRouter ranking headers (ignored by other providers).
    if "openrouter.ai" in provider["url"]:
        headers["HTTP-Referer"] = os.environ.get("APP_URL", "https://factory-scanner.local")
        headers["X-Title"] = "Factory Scanner"
    return requests.post(provider["url"], headers=headers, json=payload, timeout=60)


# ── HELPERS ───────────────────────────────────────────────────────────────────

def extract_json(text: str) -> dict:
    cleaned = text.strip().replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", cleaned, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(0))
            except json.JSONDecodeError:
                pass
        # Last resort: response was truncated mid-output (hit token limit).
        # Salvage the "table" array up to the last fully-closed row.
        salvaged = _salvage_truncated_table(cleaned)
        if salvaged is not None:
            log.warning("Recovered truncated JSON: %d complete row(s).", len(salvaged))
            return {"table": salvaged}
        raise


def _salvage_truncated_table(text: str) -> list | None:
    """
    Recover rows from a truncated `{"table":[[...],[...],[...` response.
    Returns the list of fully-closed inner rows, or None if unrecoverable.
    """
    start = text.find('"table"')
    if start == -1:
        return None
    bracket = text.find("[", start)
    if bracket == -1:
        return None
    body = text[bracket + 1:]  # inside the outer table array
    rows, depth, buf, in_str, esc = [], 0, [], False, False
    for ch in body:
        buf.append(ch)
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
        elif ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:  # one inner row just closed
                frag = "".join(buf).strip().lstrip(",").strip()
                try:
                    rows.append(json.loads(frag))
                except json.JSONDecodeError:
                    pass
                buf = []
            elif depth < 0:  # reached the outer closing bracket
                break
    return rows or None


def _coerce_to_table(parsed: dict) -> list | None:
    """
    Try every key/format the model might return and produce a 2D list.
    Handles: {table:[...]}, {rows:[{label,values}]}, {data:[...]}, {sheet:[...]},
             {rows:[[...]]}, top-level list, etc.
    """
    # Direct 2D array keys
    for key in ("table", "data", "sheet", "cells", "matrix", "grid"):
        val = parsed.get(key)
        if isinstance(val, list) and val:
            # Already 2D?
            if isinstance(val[0], list):
                return val
            # List of dicts (old {label, values} format)?
            if isinstance(val[0], dict) and "values" in val[0]:
                rows = []
                for r in val:
                    label  = r.get("label", "")
                    values = r.get("values", [])
                    rows.append([label] + list(values))
                return rows or None

    # rows key — could be list-of-dicts or list-of-lists
    rows_val = parsed.get("rows")
    if isinstance(rows_val, list) and rows_val:
        if isinstance(rows_val[0], list):
            return rows_val
        if isinstance(rows_val[0], dict):
            rows = []
            for r in rows_val:
                label  = r.get("label", "")
                values = r.get("values", [])
                rows.append([label] + list(values))
            return rows or None

    return None


def normalize_cell(v):
    """Normalize a cell value to int, float, str, or None. Preserves all text including Gujarati."""
    if v is None:
        return None
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v == int(v) else v
    s = str(v).strip()
    if not s:
        return None
    # Try numeric conversion first
    try:
        return int(s)
    except ValueError:
        pass
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except ValueError:
        pass
    return s  # preserve text as-is (Gujarati, English, etc.)


def normalize_table(raw) -> list:
    """Validate and normalize a 2D table."""
    if not raw or not isinstance(raw, list):
        return []
    rows = []
    for row in raw:
        if isinstance(row, list):
            rows.append([normalize_cell(v) for v in row])
        elif isinstance(row, dict):
            # Gracefully handle a row that came back as a dict
            rows.append([normalize_cell(v) for v in row.values()])
        else:
            rows.append([normalize_cell(row)])
    if not rows:
        return []
    num_cols = max((len(r) for r in rows), default=0)
    if not num_cols:
        return []
    for row in rows:
        while len(row) < num_cols:
            row.append(None)
    return rows


def safe_filename(value, fallback: str = "factory-data") -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", str(value or fallback)).strip("-._")
    return cleaned or fallback


def err(message: str, status: int = 400):
    log.warning("Response %d: %s", status, message)
    return jsonify({"error": message}), status


# ── EXCEL BUILDER ─────────────────────────────────────────────────────────────

def build_excel(table) -> BytesIO | None:
    table = normalize_table(table)
    if not table:
        return None

    num_rows = len(table)
    num_cols = len(table[0])

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    total_fill   = PatternFill("solid", fgColor="E2EFDA")
    grand_fill   = PatternFill("solid", fgColor="C6EFCE")
    total_font   = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align   = Alignment(wrap_text=True, vertical="top")

    total_col = num_cols + 1  # Excel column index for row-total column

    # Write extracted table + row-total formula for every row
    for ri, row in enumerate(table, start=1):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = wrap_align

        # Row total: sum all data columns
        first = get_column_letter(1)
        last  = get_column_letter(num_cols)
        tc = ws.cell(row=ri, column=total_col,
                     value=f"=SUM({first}{ri}:{last}{ri})")
        tc.font = total_font; tc.fill = total_fill; tc.alignment = center_align

    # Column-total row at the bottom (all rows, all cols)
    tr = num_rows + 1
    for ci in range(1, num_cols + 1):
        cl = get_column_letter(ci)
        tc = ws.cell(row=tr, column=ci, value=f"=SUM({cl}1:{cl}{num_rows})")
        tc.font = total_font; tc.fill = total_fill; tc.alignment = center_align

    # Grand total (sum of total column, all data rows)
    gtcl = get_column_letter(total_col)
    gc = ws.cell(row=tr, column=total_col,
                 value=f"=SUM({gtcl}1:{gtcl}{num_rows})")
    gc.font = Font(bold=True); gc.fill = grand_fill; gc.alignment = center_align

    # Auto column widths
    for ci in range(1, total_col + 1):
        max_len = 0
        for row in table:
            if ci <= len(row):
                v = row[ci - 1]
                if v is not None:
                    max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 8), 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── SECURITY HEADERS ──────────────────────────────────────────────────────────

@app.after_request
def add_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"]         = "DENY"
    response.headers["Referrer-Policy"]          = "strict-origin-when-cross-origin"
    return response


# ── ROUTES ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(PUBLIC_DIR, "index.html")


@app.route("/health")
def health():
    return jsonify({
        "ok": True,
        "providers": [{"name": p["name"], "models": p["models"]} for p in PROVIDERS],
        "primary": {"provider": PROVIDERS[0]["name"], "model": PROVIDERS[0]["models"][0]} if ATTEMPTS else None,
    })


@app.route("/extract", methods=["POST"])
def extract():
    if not ATTEMPTS:
        return err("No LLM provider is configured on the server.", 500)

    data       = request.get_json(silent=True) or {}
    image_b64  = data.get("image_base64")
    media_type = data.get("media_type", "image/jpeg")

    if not image_b64:
        return err("No image provided.", 400)
    if not isinstance(image_b64, str):
        return err("image_base64 must be a string.", 400)
    if len(image_b64.encode()) > 6 * 1024 * 1024:
        return err("Image payload too large. Reduce photo size or resolution.", 413)

    log.info("Extracting: media=%s", media_type)

    processed_b64, processed_type = preprocess_image(image_b64)

    rate_limited_any = False
    fail_msg, fail_code = "Extraction failed. Please use a clearer image.", 500

    # Try each (provider, model) in order; advance on 429 or failure. This spans
    # providers (e.g. Cerebras → OpenRouter) so one throttled pool never fails us.
    for provider, model in ATTEMPTS:
        tag = f"{provider['name']}/{model}"
        try:
            resp = call_llm(processed_b64, processed_type, provider, model)
        except requests.Timeout:
            log.warning("Timeout on %s; trying next.", tag)
            fail_msg, fail_code = "Request to the model provider timed out. Please try again.", 504
            continue
        except requests.RequestException as exc:
            log.warning("Network error on %s: %s; trying next.", tag, exc)
            fail_msg, fail_code = f"Could not reach the model provider: {exc}", 502
            continue

        if resp.status_code == 429:
            log.warning("Rate limit (429) on %s; trying next. %s", tag, resp.text[:200])
            rate_limited_any = True
            continue
        if resp.status_code != 200:
            log.error("Provider error %d on %s: %s", resp.status_code, tag, resp.text[:300])
            fail_msg, fail_code = f"Model provider returned error {resp.status_code}. Please try again.", 502
            continue

        text = None
        try:
            choice = resp.json()["choices"][0]
            text   = choice["message"]["content"]
            if choice.get("finish_reason") == "length":
                log.warning("%s hit token limit — response may be truncated.", tag)
            parsed = extract_json(text)
        except Exception as exc:
            log.error("JSON parse error on %s: %s | raw: %.200s", tag, exc, text or "")
            fail_msg, fail_code = "Could not parse the AI response. Try a clearer image.", 500
            continue

        raw_table = _coerce_to_table(parsed)
        table = normalize_table(raw_table) if raw_table is not None else []
        if not table:
            log.warning("No structured table from %s. Keys: %s", tag,
                        list(parsed.keys()) if isinstance(parsed, dict) else type(parsed).__name__)
            fail_msg, fail_code = (
                "No structured data found in the image. "
                "Make sure the image contains a table, spreadsheet, or form.", 422)
            continue

        log.info("Extracted via %s: %d rows × %d cols.", tag, len(table), len(table[0]))
        return jsonify({"table": table, "model": model, "provider": provider["name"]})

    if rate_limited_any:
        return err("All models are rate-limited right now. Wait about a minute and try again.", 429)
    return err(fail_msg, fail_code)


@app.route("/download-excel", methods=["POST"])
def download_excel():
    data   = request.get_json(silent=True) or {}
    output = build_excel(data.get("table"))
    if output is None:
        return err("No table data provided.", 400)

    filename = safe_filename(data.get("filename"), "factory-data")
    log.info("Serving Excel: %s.xlsx", filename)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{filename}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/download-all-excels", methods=["POST"])
def download_all_excels():
    data      = request.get_json(silent=True) or {}
    documents = data.get("documents") or []
    ready     = [d for d in documents if d.get("table")]
    if not ready:
        return err("No completed documents provided.", 400)

    buf  = BytesIO()
    used: set[str] = set()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for i, doc in enumerate(ready, start=1):
            wb = build_excel(doc.get("table"))
            if wb is None:
                continue
            base  = safe_filename(doc.get("filename"), f"image-{i}")
            fname = f"{base}.xlsx"
            n = 2
            while fname.lower() in used:
                fname = f"{base}-{n}.xlsx"
                n += 1
            used.add(fname.lower())
            zf.writestr(fname, wb.getvalue())

    buf.seek(0)
    log.info("Serving ZIP with %d Excel file(s).", len(ready))
    return send_file(
        buf,
        as_attachment=True,
        download_name="factory-scanner-excels.zip",
        mimetype="application/zip",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    log.info("Starting Factory Scanner (dev) on http://localhost:%d", port)
    app.run(host="0.0.0.0", debug=False, port=port)
